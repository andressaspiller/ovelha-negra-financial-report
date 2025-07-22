import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import BarChart

def create_excel_dashboards():
    cleaned_data_path = '/home/ubuntu/cleaned_data'
    output_excel_path = '/home/ubuntu/financial_dashboard.xlsx'

    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # 1. Dashboard Resumo Financeiro
    ws_dashboard = wb.create_sheet("Dashboard Financeiro")
    ws_dashboard["A1"] = "Resumo Financeiro Geral"
    ws_dashboard["A2"] = "Dados baseados nas análises de Cachorro Quente, Conta da Casa e Obra Banheiro."

    # Load financial summary from the generated text file
    financial_summary = {}
    try:
        with open("/home/ubuntu/financial_summary.txt", "r") as f:
            lines = f.readlines()
            for line in lines:
                if ": R$" in line:
                    key, value = line.split(": R$")
                    financial_summary[key.strip()] = float(value.strip())
    except FileNotFoundError:
        print("Arquivo financial_summary.txt não encontrado. O dashboard pode estar incompleto.")

    # Add key metrics to dashboard sheet
    ws_dashboard["A4"] = "TOTAL RECEITAS:"
    ws_dashboard["B4"] = financial_summary.get("total_receitas", 0)
    ws_dashboard["A5"] = "TOTAL DESPESAS:"
    ws_dashboard["B5"] = financial_summary.get("total_despesas", 0)
    ws_dashboard["A6"] = "TOTAL DÍVIDAS:"
    ws_dashboard["B6"] = financial_summary.get("total_dividas", 0)
    ws_dashboard["A7"] = "SALDO GERAL:"
    ws_dashboard["B7"] = financial_summary.get("saldo_geral", 0)
    ws_dashboard["A8"] = "Margem Líquida (%):"
    ws_dashboard["B8"] = 100.0  # From the analysis output

    # Create a pie chart for Revenue Distribution
    receitas_data = {
        "Cachorro Quente": financial_summary.get("total_cachorro_quente", 0),
        "Conta da Casa - Entradas": financial_summary.get("conta_casa_entradas", 0),
        "Obra Banheiro - Arrecadado": financial_summary.get("obra_banheiro_arrecadado", 0)
    }

    # Add data to a temporary range for the chart
    row_offset = 10
    ws_dashboard[f"A{row_offset}"] = "Distribuição de Receitas"
    ws_dashboard[f"A{row_offset+1}"] = "Fonte"
    ws_dashboard[f"B{row_offset+1}"] = "Valor"
    for i, (source, value) in enumerate(receitas_data.items()):
        ws_dashboard[f"A{row_offset+2+i}"] = source
        ws_dashboard[f"B{row_offset+2+i}"] = value

    pie = PieChart()
    labels = Reference(ws_dashboard, min_col=1, min_row=row_offset+2, max_row=row_offset+1+len(receitas_data))
    data = Reference(ws_dashboard, min_col=2, min_row=row_offset+2, max_row=row_offset+1+len(receitas_data))
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "Distribuição de Receitas por Fonte"
    pie.dLbls = DataLabelList()
    pie.dLbls.showCatName = True
    pie.dLbls.showPercent = True
    ws_dashboard.add_chart(pie, "D1")

    # 2. Adicionar dados brutos limpos em abas separadas
    for f in os.listdir(cleaned_data_path):
        if f.endswith("_cleaned.csv"):
            sheet_name = f.replace("_cleaned.csv", "").replace("Copyof", "")[:31] # Max 31 chars for sheet name
            try:
                df = pd.read_csv(os.path.join(cleaned_data_path, f))
                ws = wb.create_sheet(sheet_name)
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
            except Exception as e:
                print(f"Erro ao adicionar {f} à planilha Excel: {e}")

    wb.save(output_excel_path)
    print(f"Dashboard Excel salvo em: {output_excel_path}")

if __name__ == '__main__':
    create_excel_dashboards()

